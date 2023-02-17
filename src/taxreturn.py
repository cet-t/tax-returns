import openpyxl
sheets = [
    'hiroko_med',
    'hiroko_nur',
    'takashi_med',
    'takashi_nur'
]

max_row = [
    30,
    39,
    64,
    75
]

titles = [
    'hiroko medical',
    'hiroko nursing',
    'takashi medical',
    'takashi nursing',
    'nursing',
]

path = 'data/tax_return2.xlsx'


def tax_culc(_name: str, _sheet: int, _max: int) -> None:
    file = openpyxl.load_workbook(path)
    sheet = file[_sheet]
    sums = []
    for i in range(1, _max, 1):
        datas = sheet.cell(i, 2).value
        sums.append(datas)
    print(f'{_name}: {sum(sums)}')

    file.close()


def tax_sum(_sheet: int, _max: int) -> int:
    file = openpyxl.load_workbook(path)
    sheet = file[_sheet]
    sums = []
    for i in range(1, _max, 1):
        datas = sheet.cell(i, 2).value
        sums.append(datas)
    file.close()
    return sum(sums)


sums: list[int]
sums = [
    tax_sum(sheets[0], max_row[0]),
    tax_sum(sheets[1], max_row[1]),
    tax_sum(sheets[2], max_row[2]),
    tax_sum(sheets[3], max_row[3]),
]

alls: str = [
    f'hiroko-medical: {sums[0]}',  # 医1
    f'hiroko-nursing: {sums[1]}',  # 介1
    f'hiroko-subtotal: {sums[0]+sums[1]}',  # 合計1
    '',
    f'takashi-medical: {sums[2]}',  # 医2
    f'takashi-nursing: {sums[3]}',  # 介2
    f'takashi-subtotal: {sums[2]+sums[3]}',  # 合計2
    '',
    f'medical-subtotal: {sums[0]+sums[2]}',  # 医合計
    f'nursing-subtotal: {sums[1]+sums[3]}',  # 介合計
    f'total: {sum(sums)}',
]
if __name__ == '__main__':
    import openpyxl

    for i in alls:
        print(i)
